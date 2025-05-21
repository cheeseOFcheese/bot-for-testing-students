import os
import datetime
from openpyxl import Workbook, load_workbook
from telegram import Update
from telegram.ext import Updater, CommandHandler, MessageHandler, Filters, ConversationHandler, CallbackContext

# States for ConversationHandler
ASK_NAME, ASK_QUESTION = range(2)

# Define your questions and answers
test_questions = [
    {
        "question": "1. Что такое полигональная модель в 3D-графике?\nA) Модель, построенная на основе полигонов\nB) Модель, использующая только кривые\nC) Текстурированная модель\nВведите A, B или C.",
        "answer": "A"
    },
    {
        "question": "2. Какой формат файла чаще всего используется для переноса 3D-сцен между приложениями?\nA) .png\nB) .obj\nC) .mp4\nВведите A, B или C.",
        "answer": "B"
    },
    {
        "question": "3. Что такое рендеринг в контексте CG-технологий?\nA) Процесс моделирования\nB) Процесс визуализации сцены\nC) Процесс захвата движений
Введите A, B или C.",
        "answer": "B"
    },
    {
        "question": "4. В ивент-индустрии для чего применяют LED-стены и проекционные экраны?\nA) Для озвучивания сцены\nB) Для визуального оформления и отображения 3D-контента\nC) Для охлаждения оборудования
Введите A, B или C.",
        "answer": "B"
    },
    {
        "question": "5. Что означает термин 'motion capture'?\nA) Захват движения актера для последующей анимации\nB) Запись звука на площадке\nC) Монтаж видео
Введите A, B или C.",
        "answer": "A"
    }
]

EXCEL_FILE = 'results.xlsx'

# Ensure Excel file exists and has header
def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(["Timestamp", "Name", "Score", "Max"])
        wb.save(EXCEL_FILE)

# Append a row to Excel
def save_result(name: str, score: int, max_score: int):
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    timestamp = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    ws.append([timestamp, name, score, max_score])
    wb.save(EXCEL_FILE)

# Start command handler
def start(update: Update, context: CallbackContext) -> int:
    update.message.reply_text(
        "Привет! Я бот для тестирования по теме 3D-графики и CG технологий в ивент-индустрии. Как тебя зовут?"
    )
    return ASK_NAME

# Name handler
def ask_questions(update: Update, context: CallbackContext) -> int:
    name = update.message.text.strip()
    context.user_data['name'] = name
    context.user_data['score'] = 0
    context.user_data['q_index'] = 0
    update.message.reply_text(f"Приятно познакомиться, {name}! Начинаем тест.\n{test_questions[0]['question']}")
    return ASK_QUESTION

# Question handler
def handle_answer(update: Update, context: CallbackContext) -> int:
    text = update.message.text.strip().upper()
    idx = context.user_data['q_index']
    correct = test_questions[idx]['answer']
    if text == correct:
        context.user_data['score'] += 1
    idx += 1
    if idx < len(test_questions):
        context.user_data['q_index'] = idx
        update.message.reply_text(test_questions[idx]['question'])
        return ASK_QUESTION
    # Test finished
    name = context.user_data['name']
    score = context.user_data['score']
    max_score = len(test_questions)
    update.message.reply_text(
        f"Тест завершён! {name}, твой результат: {score} из {max_score}."
    )
    save_result(name, score, max_score)
    return ConversationHandler.END

# Cancel handler
def cancel(update: Update, context: CallbackContext) -> int:
    update.message.reply_text('Тест прерван.')
    return ConversationHandler.END

if __name__ == '__main__':
    init_excel()
    TOKEN = os.getenv('TELEGRAM_BOT_TOKEN')  # Установите переменную окружения перед запуском
    if not TOKEN:
        print("Error: TELEGRAM_BOT_TOKEN не задан")
        exit(1)

    updater = Updater(TOKEN)
    dp = updater.dispatcher

    conv = ConversationHandler(
        entry_points=[CommandHandler('start', start)],
        states={
            ASK_NAME: [MessageHandler(Filters.text & ~Filters.command, ask_questions)],
            ASK_QUESTION: [MessageHandler(Filters.text & ~Filters.command, handle_answer)],
        },
        fallbacks=[CommandHandler('cancel', cancel)]
    )
    dp.add_handler(conv)

    print("Бот запущен...")
    updater.start_polling()
    updater.idle()
